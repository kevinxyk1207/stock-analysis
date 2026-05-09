"""
综合日报生成器 — Excel多sheet格式
Sheet: 市场概览 | 纸上持仓 | 长线Top10 | 10d短线 | 技术信号 | 关键价位 | 洞察 | 建议 | 早期关注 | 精筛池
"""
import sys, os, json
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from enhanced_fetcher import load_cache_data, load_stock_name_map, EnhancedStockFetcher
from b1_selector import B1Selector, B1Config

REPORT_DIR = "reports"

# 样式
HEADER_FILL = PatternFill(start_color="FF2c3e50", end_color="FF2c3e50", fill_type="solid")
HEADER_FONT = Font(color="FFFFFFFF", bold=True, size=11)
GREEN_FILL = PatternFill(start_color="FFd4edda", end_color="FFd4edda", fill_type="solid")
RED_FILL = PatternFill(start_color="FFf8d7da", end_color="FFf8d7da", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFfff3cd", end_color="FFfff3cd", fill_type="solid")
BOLD = Font(bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)


def get_market_regime(stock_data):
    close_dict = {s: pd.Series(df["close"].values) for s, df in stock_data.items()}
    mc = pd.DataFrame(close_dict).mean(axis=1)
    mma60 = mc.rolling(60, min_periods=60).mean()
    if len(mma60) > 0 and pd.notna(mma60.iloc[-1]):
        return mc.iloc[-1] > mma60.iloc[-1], round(float(mc.iloc[-1]), 2), round(float(mma60.iloc[-1]), 2)
    return True, 0, 0


def load_portfolio():
    pf_path = os.path.join(REPORT_DIR, "paper_portfolio.json")
    if os.path.exists(pf_path):
        with open(pf_path, "r", encoding="utf-8") as f:
            return json.load(f)
    return None


def analyze_stock(sym, df, selector, fin_data=None, deep_insights=None):
    try:
        fin = fin_data.get(sym, {}) if fin_data else {}
        di = deep_insights.get(sym, {}) if deep_insights else {}
        prepared = selector.prepare_data(df)
        cond = selector.check_b1_conditions(prepared, date_idx=-1)
        close = float(prepared["close"].iloc[-1])
        s10 = selector._calculate_score(cond, "short")
        s60 = selector._calculate_score_long(cond)
        rsi = round(cond.get("rsi", 50), 1)
        j_val = round(cond.get("j_value", 50), 1)
        zr = round(cond.get("zxdkx_ratio", 1.0), 3)
        ma_align = int(cond.get("ma_alignment", 0))
        trend = round(cond.get("trend_strength", 0), 1)
        macd_q = round(cond.get("macd_quality", 0), 1)
        vol_trend = round(cond.get("vol_trend_ratio", 1.0), 3)

        rets = {}
        for period, label in [(5, "5d"), (10, "10d"), (20, "20d"), (60, "60d")]:
            if len(prepared) > period:
                rets[label] = round((close - prepared["close"].iloc[-period - 1]) / prepared["close"].iloc[-period - 1] * 100, 1)

        high_20 = round(float(prepared["close"].iloc[-20:].max()), 2)
        low_20 = round(float(prepared["close"].iloc[-20:].min()), 2)
        high_60 = round(float(prepared["close"].iloc[-60:].max()), 2)
        at_high = close >= high_60 * 0.98

        mas = {}
        for p in [5, 10, 20, 60]:
            mas[p] = round(float(prepared["close"].rolling(p).mean().iloc[-1]), 2)
        zxdkx_val = round((mas[5] + mas[20] + mas[60] + mas.get(40, mas[20])) / 4, 2)

        tr = pd.concat([
            (prepared["high"] - prepared["low"]).abs(),
            (prepared["high"] - prepared["close"].shift(1)).abs(),
            (prepared["low"] - prepared["close"].shift(1)).abs(),
        ], axis=1).max(axis=1)
        atr = round(float(tr.rolling(14).mean().iloc[-1]), 2)

        # 风险信号（仅真正的恶化指标）
        risks = []
        if rsi < 30: risks.append("RSI超卖")
        if zr < 0.98: risks.append("跌破知行线")
        if trend < -3: risks.append("趋势恶化")
        ret60 = rets.get("60d")
        if ret60 is not None and ret60 < -20: risks.append("60日跌幅过大")

        # 综合星级评定（★=20分，满分100）
        star_score = 0

        # 1) 60d技术得分 (0-30分)
        if s60 >= 90: star_score += 30
        elif s60 >= 85: star_score += 25
        elif s60 >= 80: star_score += 20
        elif s60 >= 75: star_score += 15
        elif s60 >= 70: star_score += 10
        else: star_score += 5

        # 2) 风险扣分 (-20~0)
        star_score -= len(risks) * 8

        # 3) MACD动量 (0-20分)
        if macd_q > 150: star_score += 20
        elif macd_q > 100: star_score += 15
        elif macd_q > 50: star_score += 10
        elif macd_q > 0: star_score += 5

        # 4) 均线结构 (0-15分)
        if ma_align == 3: star_score += 15
        elif ma_align == 2: star_score += 8

        # 5) 趋势强度 (0-10分)
        if trend > 8: star_score += 10
        elif trend > 5: star_score += 7
        elif trend > 2: star_score += 4
        elif trend > 0: star_score += 2

        # 6) RSI位置 (0-5分)
        if 50 <= rsi <= 70: star_score += 5
        elif 70 < rsi <= 80: star_score += 3
        elif rsi > 80: star_score += 0  # 超买不加分不减分

        # 7) 量能配合 (0-10分)
        if vol_trend > 1.3: star_score += 10
        elif vol_trend > 1.1: star_score += 7
        elif vol_trend > 0.9: star_score += 3

        # 星级映射
        # 8) 基本面加分 (0-25分)
        fund_score = 0
        fund_details = []
        rev_g = fin.get("营收增速", 0)
        prf_g = fin.get("利润增速", 0)
        margin = fin.get("毛利率", 0)
        roe = fin.get("ROE", 0)
        if rev_g and rev_g > 50: fund_score += 6; fund_details.append(f"营收+{rev_g:.0f}%")
        elif rev_g and rev_g > 30: fund_score += 4; fund_details.append(f"营收+{rev_g:.0f}%")
        elif rev_g and rev_g > 20: fund_score += 2; fund_details.append(f"营收+{rev_g:.0f}%")
        elif rev_g and rev_g > 10: fund_score += 1
        if margin and margin > 50: fund_score += 5; fund_details.append(f"毛利{margin:.0f}%")
        elif margin and margin > 40: fund_score += 3
        elif margin and margin > 30: fund_score += 1
        if roe and roe > 15: fund_score += 4
        elif roe and roe > 10: fund_score += 2
        # 利润质量：利润增速远超营收增速→扣分(不可持续)
        if rev_g and prf_g and prf_g > rev_g * 5 and rev_g > 0:
            fund_score -= 3; fund_details.append("利润质量存疑")
        # 深层洞察调整
        if di:
            verdict_ov = di.get("verdict_override", "")
            if "核心标的" in verdict_ov or "首选" in verdict_ov: fund_score += 3
            if "风险最多" in verdict_ov or "不宜重仓" in verdict_ov: fund_score -= 3

        # 合并得分
        total_score = star_score + max(-10, min(25, fund_score))
        fund_label = f"(基本面上修{fund_score:+d})" if fund_score != 0 else ""

        if total_score >= 90: verdict = "★★★★★"
        elif total_score >= 75: verdict = "★★★★"
        elif total_score >= 60: verdict = "★★★"
        elif total_score >= 45: verdict = "★★"
        else: verdict = "★"

        return {
            "symbol": sym, "close": close, "score_short": round(s10,1), "score_long": round(s60,1),
            "rsi": rsi, "j_val": j_val, "zxdkx_ratio": zr, "ma_alignment": ma_align,
            "trend": trend, "macd_quality": macd_q, "vol_trend": vol_trend,
            "rets": rets, "high_20": high_20, "low_20": low_20, "high_60": high_60,
            "at_high": at_high, "mas": mas, "zxdkx_val": zxdkx_val,
            "atr": atr, "risks": risks, "verdict": verdict,
        }
    except Exception as e:
        return {"symbol": sym, "error": str(e)}


def generate_report():
    t0 = datetime.now()
    fetcher = EnhancedStockFetcher()
    hs300 = fetcher.get_hs300_stocks()
    hs300_set = set(str(s).zfill(6) for s in hs300["code"].tolist())

    stock_data_raw = load_cache_data(min_rows=200, common_range=False)
    stock_data = {k: v for k, v in stock_data_raw.items() if str(k).zfill(6) in hs300_set}
    last_date = max(df.index[-1] for df in stock_data.values())
    is_bull, mc, ma = get_market_regime(stock_data)

    config = B1Config(j_threshold=10.0, j_q_threshold=0.30, kdj_n=9,
        zx_m1=5, zx_m2=20, zx_m3=40, zx_m4=60, zxdq_span=10,
        wma_short=5, wma_mid=10, wma_long=15, max_vol_lookback=20)
    selector = B1Selector(config)
    name_map = load_stock_name_map()

    ind_map = {}
    if os.path.exists("data_cache/industry_map.json"):
        with open("data_cache/industry_map.json", encoding="utf-8") as f:
            ind_map = json.load(f)

    # ── 深层洞察缓存 ──
    deep_insights = {}
    if os.path.exists("data_cache/fundamental_insights.json"):
        with open("data_cache/fundamental_insights.json", encoding="utf-8") as f:
            di = json.load(f)
            deep_insights = di.get("stocks", {})

    # ── 基本面数据（akshare Q1季报）──
    fin_data = {}
    try:
        import akshare as ak
        df_fin = ak.stock_yjbb_em(date="20260331")
        for _, row in df_fin.iterrows():
            code = str(row["股票代码"]).zfill(6)
            fin_data[code] = {
                "营收": row.get("营业总收入-营业总收入", 0),
                "营收增速": row.get("营业总收入-同比增长", 0),
                "净利润": row.get("净利润-净利润", 0),
                "利润增速": row.get("净利润-同比增长", 0),
                "毛利率": row.get("销售毛利率", 0),
                "ROE": row.get("净资产收益率", 0),
                "每股收益": row.get("每股收益", 0),
            }
    except Exception as e:
        print(f"  [警告] 基本面数据获取失败: {e}")

    # 长线评分
    results = []
    for sym, df in stock_data.items():
        sym = str(sym).zfill(6)
        try:
            if df.empty or len(df) < 60: continue
            prepared = selector.prepare_data(df)
            cond = selector.check_b1_conditions(prepared, date_idx=-1)
            score = selector._calculate_score_long(cond)
            results.append((sym, df, score))
        except: continue
    results.sort(key=lambda x: x[2], reverse=True)
    top10 = results[:10]

    # 短线评分
    results_10d = []
    for sym, df in stock_data.items():
        sym = str(sym).zfill(6)
        try:
            if df.empty or len(df) < 60: continue
            prepared = selector.prepare_data(df)
            cond = selector.check_b1_conditions(prepared, date_idx=-1)
            score = selector._calculate_score(cond, "short")
            results_10d.append((sym, df, score))
        except: continue
    results_10d.sort(key=lambda x: x[2], reverse=True)
    top10_10d = results_10d[:10]

    # 纸上持仓
    pf = load_portfolio()
    pf_rows = []
    if pf:
        for s in pf.get("stocks", []):
            for f in os.listdir(fetcher.cache_dir):
                if not f.endswith(".csv"): continue
                if f[:-4] == s["symbol"] or str(f[:-4]).zfill(6) == str(s["symbol"]).zfill(6):
                    tdf = pd.read_csv(os.path.join(fetcher.cache_dir, f), index_col=0, parse_dates=True)
                    if not tdf.empty:
                        cp = float(tdf["close"].iloc[-1])
                        pnl = (cp - s["entry_price"]) / s["entry_price"] * 100
                        pf_rows.append({
                            "代码": s["symbol"], "名称": s["name"],
                            "买入价": s["entry_price"], "现价": round(cp, 2),
                            "盈亏%": round(pnl, 2), "入场得分": s["score"],
                        })
                    break

    # ── 创建 Excel ──
    today_str = last_date.strftime("%Y%m%d")
    path = os.path.join(REPORT_DIR, f"daily_report_{today_str}.xlsx")
    os.makedirs(REPORT_DIR, exist_ok=True)

    wb = pd.ExcelWriter(path, engine="openpyxl")

    # Sheet 1: 市场概览
    overview = pd.DataFrame([
        ["数据日期", str(last_date.date())],
        ["生成时间", t0.strftime("%Y-%m-%d %H:%M")],
        ["市场状态", "牛市" if is_bull else "熊市"],
        ["HS300均价", f"{mc:.2f}"],
        ["MA60", f"{ma:.2f}"],
        ["选股范围", f"沪深300 ({len(stock_data)}只)"],
        ["操作建议", "正常选股，满仓操作" if is_bull else "建议空仓观望"],
    ], columns=["指标", "数值"])
    overview.to_excel(wb, sheet_name="市场概览", index=False)

    # Sheet 1.5: 实盘持仓（真实持仓跟踪）
    holdings_path = os.path.join("data_cache", "my_holdings.json")
    if os.path.exists(holdings_path):
        with open(holdings_path, encoding="utf-8") as f:
            holdings = json.load(f).get("stocks", [])
        if holdings:
            h_rows = []
            for s in holdings:
                sym = s["symbol"]
                entry_p = s.get("entry_price")
                entry_d = s.get("entry_date", "")
                # 当前价+技术面
                current_p = None; sell_sig = "—"; rsi_v = None; macd_q = None; trend_v = None
                for fname in [f"{sym}.csv", f"{sym.lstrip('0')}.csv"]:
                    fp = os.path.join("data_cache", fname)
                    if os.path.exists(fp):
                        try:
                            tdf = pd.read_csv(fp, index_col=0, parse_dates=True)
                            if not tdf.empty and len(tdf) >= 100:
                                pdf = selector.prepare_data(tdf)
                                cond = selector.check_b1_conditions(pdf, date_idx=-1)
                                current_p = round(float(pdf["close"].iloc[-1]), 2)
                                rsi_v = round(cond.get("rsi", 50), 1)
                                macd_q = round(cond.get("macd_quality", 0), 1)
                                trend_v = round(cond.get("trend_strength", 0), 1)
                                # v3卖出信号
                                weekly = tdf['close'].resample('W-FRI').last().dropna()
                                wh = tdf['high'].resample('W-FRI').max()
                                wv = tdf['volume'].resample('W-FRI').sum()
                                if len(weekly) >= 14:
                                    wc = weekly.iloc[-1]; wc_p = weekly.iloc[-2]
                                    rh12 = wh.iloc[-12:].max()
                                    pv12 = wv.iloc[-12:].max() * weekly.iloc[-12:].max()/1e8
                                    cv = wv.iloc[-1]*wc/1e8; cv_p = wv.iloc[-2]*wc_p/1e8
                                    ma5w = weekly.iloc[-5:].mean()
                                    delta = weekly.iloc[-14:].diff()
                                    g = delta.clip(lower=0).mean(); l = (-delta).clip(lower=0).mean()
                                    rsi_w = 100-100/(1+g/(l+1e-9)) if l > 0 else 100
                                    near = wc >= rh12*0.97
                                    div2 = near and cv < pv12*0.5 and cv_p < pv12*0.5
                                    bma = wc < ma5w*0.95
                                    hot = rsi_w > 82 and wc > ma5w*1.15
                                    # 简化判断：不要求连续两周（可能只有最新一周的数据）
                                    single_div = near and cv < pv12 * 0.5
                                    if div2: sell_sig = "建议卖出"
                                    elif bma: sell_sig = "建议卖出"
                                    elif hot: sell_sig = "建议卖出"
                                    else: sell_sig = "持有"
                        except Exception: pass
                        break
                if entry_p and current_p:
                    pnl = (current_p/entry_p - 1) * 100
                    pnl_str = f"{pnl:+.1f}%"
                else:
                    pnl_str = "—"
                h_rows.append({
                    "代码": sym, "名称": s.get("name",""),
                    "成本": f"{entry_p:.2f}" if entry_p else "—",
                    "现价": f"{current_p:.2f}" if current_p else "—",
                    "盈亏": pnl_str,
                    "RSI": f"{rsi_v:.0f}" if rsi_v else "—",
                    "MACD": f"{macd_q:.0f}" if macd_q else "—",
                    "趋势": f"{trend_v:+.1f}%" if trend_v else "—",
                    "卖出": sell_sig,
                })
            pd.DataFrame(h_rows).to_excel(wb, sheet_name="实盘持仓", index=False)

    # Sheet 2: 纸上持仓
    if pf_rows:
        pf_df = pd.DataFrame(pf_rows)
        pf_df.to_excel(wb, sheet_name="纸上持仓", index=False)
    else:
        pd.DataFrame([["暂无持仓"]]).to_excel(wb, sheet_name="纸上持仓", index=False)

    # Sheet 3: 长线 Top10 (20d持有)
    summary_rows = []
    for i, (sym, df, score) in enumerate(top10, 1):
        name = name_map.get(sym, "?")
        a = analyze_stock(sym, df, selector, fin_data, deep_insights)
        rets = a.get("rets", {})
        fin = fin_data.get(sym, {})
        summary_rows.append({
            "排名": i, "代码": sym, "名称": name,
            "收盘": a.get("close", 0), "长线得分": a.get("score_long", 0),
            "评级": a.get("verdict", "?"),
            "风险": ", ".join(a.get("risks", [])),
            "zxdkx比": a.get("zxdkx_ratio", 0), "RSI": a.get("rsi", 0),
            "趋势%": a.get("trend", 0),
            "5日%": rets.get("5d", 0), "20日%": rets.get("20d", 0),
            "Q1营收(亿)": round(fin.get("营收", 0) / 1e8, 1) if fin.get("营收") else None,
            "营收增速%": round(fin.get("营收增速", 0), 1) if fin.get("营收增速") else None,
            "Q1净利(亿)": round(fin.get("净利润", 0) / 1e8, 1) if fin.get("净利润") else None,
            "利润增速%": round(fin.get("利润增速", 0), 1) if fin.get("利润增速") else None,
            "毛利率%": round(fin.get("毛利率", 0), 1) if fin.get("毛利率") else None,
            "ATR%": round(a.get("atr", 0) / a.get("close", 1) * 100, 1) if a.get("close", 0) > 0 else 0,
        })
    df_sum = pd.DataFrame(summary_rows)
    df_sum.to_excel(wb, sheet_name="长线 Top10", index=False)

    # Sheet 3.5: 10d 短线 Top10 参考
    short_rows = []
    for i, (sym, df, score) in enumerate(top10_10d, 1):
        name = name_map.get(sym, "?")
        a = analyze_stock(sym, df, selector, fin_data, deep_insights)
        rets = a.get("rets", {})
        short_rows.append({
            "排名": i, "代码": sym, "名称": name,
            "收盘": a.get("close", 0), "短线得分": a.get("score_short", 0),
            "长线得分": a.get("score_long", 0),
            "zxdkx比": a.get("zxdkx_ratio", 0), "RSI": a.get("rsi", 0),
            "J值": a.get("j_val", 0),
            "5日%": rets.get("5d", 0), "20日%": rets.get("20d", 0),
            "评级": a.get("verdict", "?"),
        })
    pd.DataFrame(short_rows).to_excel(wb, sheet_name="短线参考", index=False)

    # Sheet 4: 技术深度（信号+价位合并，避免重复调用analyze_stock）
    tech_deep = []
    for i, (sym, df, score) in enumerate(top10, 1):
        name = name_map.get(sym, "?")
        a = analyze_stock(sym, df, selector, fin_data, deep_insights)
        rets = a.get("rets", {})
        mas = a.get("mas", {})
        c = a.get("close", 1)
        tech_deep.append({
            "排名": i, "代码": sym, "名称": name,
            "zxdkx比": a.get("zxdkx_ratio", 0), "RSI": a.get("rsi", 0),
            "J值": a.get("j_val", 0), "均线": f"{a.get('ma_alignment',0)}/3",
            "趋势": f"{a.get('trend',0):+.1f}%", "MACD质量": a.get("macd_quality", 0),
            "5日%": rets.get("5d", 0), "20日%": rets.get("20d", 0),
            "MA5": mas.get(5, 0), "MA20": mas.get(20, 0),
            "知行线": a.get("zxdkx_val", 0), "MA60": mas.get(60, 0),
            "ATR": a.get("atr", 0), "新高?": "是" if a.get("at_high") else "否",
        })
    pd.DataFrame(tech_deep).to_excel(wb, sheet_name="技术深度", index=False)

    # Sheet 6: 深层洞察（研报/行业/风险）
    insight_rows = []
    for i, (sym, df, score) in enumerate(top10, 1):
        name = name_map.get(sym, "?")
        di = deep_insights.get(sym, {})
        if di:
            insight_rows.append({
                "排名": i, "代码": sym, "名称": name,
                "综合判断": di.get("verdict_override", ""),
                "核心亮点": " | ".join(di.get("deep_highlights", [])),
                "隐蔽风险": " | ".join(di.get("deep_risks", [])),
            })
    # Sheet 6: 综合建议（深层洞察+操作建议合并，避免重复调用analyze_stock）
    combined_rows = []
    for i, (sym, df, score) in enumerate(top10):
        name = name_map.get(sym, "?")
        a = analyze_stock(sym, df, selector, fin_data, deep_insights)
        fin = fin_data.get(sym, {})
        rets = a.get("rets", {})
        di = deep_insights.get(sym, {})
        combined_rows.append({
            "排名": i + 1, "代码": sym, "名称": name,
            "长线得分": a.get("score_long", 0), "评级": a.get("verdict", "?"),
            "风险": ", ".join(a.get("risks", [])),
            "Q1营收(亿)": round(fin.get("营收", 0) / 1e8, 1) if fin.get("营收") else None,
            "利润增速%": round(fin.get("利润增速", 0), 1) if fin.get("利润增速") else None,
            "毛利率%": round(fin.get("毛利率", 0), 1) if fin.get("毛利率") else None,
            "综合判断": di.get("verdict_override", ""),
            "核心亮点": " | ".join(di.get("deep_highlights", [])[:2]),
            "隐蔽风险": " | ".join(di.get("deep_risks", [])[:2]),
        })
    pd.DataFrame(combined_rows).to_excel(wb, sheet_name="综合建议", index=False)

    # Sheet 8: 早期关注（全市场潜力股，不限HS300）
    ew_path = os.path.join("data_cache", "early_watch.json")
    if os.path.exists(ew_path):
        with open(ew_path, encoding="utf-8") as f:
            ew_data = json.load(f)
        ew_stocks = ew_data.get("stocks", [])
        if ew_stocks:
            ew_rows = []
            for s in ew_stocks:
                sym = s["symbol"]
                name = s["name"]
                logic = s.get("core_logic", "")
                disc_date = s.get("discovered", "")
                disc_price = s.get("discovery_price")

                # 从全量缓存读取当前数据
                current_price = None
                current_score = None
                current_rank = None
                for fname in [f"{sym}.csv", f"{sym.lstrip('0')}.csv", f"{sym.zfill(6)}.csv"]:
                    fpath = os.path.join("data_cache", fname)
                    if os.path.exists(fpath):
                        try:
                            tdf = pd.read_csv(fpath, index_col=0, parse_dates=True)
                            if not tdf.empty and len(tdf) >= 60:
                                pdf = selector.prepare_data(tdf)
                                cond = selector.check_b1_conditions(pdf, date_idx=-1)
                                current_price = round(float(pdf["close"].iloc[-1]), 2)
                                current_score = round(selector._calculate_score_long(cond), 1)
                        except Exception:
                            pass
                        break

                # 计算涨幅和排名
                gain_str = ""
                rank_str = ""
                if current_price and disc_price:
                    gain = (current_price / disc_price - 1) * 100
                    gain_str = f"{gain:+.1f}%"
                if current_score:
                    # 在HS300结果中找排名(早期关注股可能不在HS300)
                    for rank, (rsym, _, sc) in enumerate(results, 1):
                        if rsym == sym or str(rsym).zfill(6) == str(sym).zfill(6):
                            current_rank = rank
                            rank_str = f"#{rank}/{len(results)}"
                            break

                ew_rows.append({
                    "代码": sym, "名称": name,
                    "发现日期": disc_date, "发现价": disc_price,
                    "现价": current_price, "涨幅": gain_str,
                    "长线得分": current_score,
                    "排名": rank_str,
                    "核心逻辑": logic,
                })
            if ew_rows:
                pd.DataFrame(ew_rows).to_excel(wb, sheet_name="早期关注", index=False)

    # Sheet 8: 精筛池（Claude优中选优 ≤5只）
    cp_path = os.path.join("data_cache", "curated_pool.json")
    if os.path.exists(cp_path):
        with open(cp_path, encoding="utf-8") as f:
            cps = json.load(f).get("stocks", [])
        if cps:
            cp_rows = []
            for s in cps:
                sym = s["symbol"]
                current_price = None; current_score = None
                for fname in [f"{sym}.csv", f"{sym.lstrip('0')}.csv"]:
                    fpath = os.path.join("data_cache", fname)
                    if os.path.exists(fpath):
                        try:
                            tdf = pd.read_csv(fpath, index_col=0, parse_dates=True)
                            if not tdf.empty and len(tdf) >= 60:
                                pdf = selector.prepare_data(tdf)
                                cond = selector.check_b1_conditions(pdf, date_idx=-1)
                                current_price = round(float(pdf["close"].iloc[-1]), 2)
                                current_score = round(selector._calculate_score_long(cond), 1)
                        except Exception: pass
                        break
                # v3 卖出信号检测
                sell_signal = ""
                if current_price and tdf is not None and len(tdf) >= 100:
                    try:
                        weekly = tdf['close'].resample('W-FRI').last().dropna()
                        wh = tdf['high'].resample('W-FRI').max()
                        wv = tdf['volume'].resample('W-FRI').sum()
                        if len(weekly) >= 14:
                            wc = weekly.iloc[-1]; wc_prev = weekly.iloc[-2]
                            rh12 = wh.iloc[-12:].max()
                            pv12 = wv.iloc[-12:].max() * weekly.iloc[-12:].max() / 1e8
                            cv = wv.iloc[-1] * wc / 1e8
                            cv_prev = wv.iloc[-2] * wc_prev / 1e8
                            ma5w = weekly.iloc[-5:].mean()
                            delta = weekly.iloc[-14:].diff()
                            g = delta.clip(lower=0).mean(); l = (-delta).clip(lower=0).mean()
                            rsi = 100-100/(1+g/(l+1e-9)) if l > 0 else 100
                            near_peak = wc >= rh12 * 0.97
                            two_week_div = near_peak and cv < pv12 * 0.5 and cv_prev < pv12 * 0.5
                            break_ma = wc < ma5w * 0.95 and cv > 1.3 * (wv.iloc[-10:-1].mean() * weekly.iloc[-10:-1].mean() / 1e8)
                            rsi_warn = rsi > 82 and wc > ma5w * 1.15
                            if two_week_div: sell_signal = "建议卖出"
                            elif break_ma: sell_signal = "建议卖出"
                            elif rsi_warn: sell_signal = "建议卖出"
                            else: sell_signal = "持有"
                    except Exception: pass

                cp_rows.append({
                    "代码": sym, "名称": s.get("name",""), "入选日": s.get("added",""),
                    "现价": f"{current_price:.2f}" if current_price else "?",
                    "卖出信号": sell_signal,
                    "核心逻辑": s.get("core_logic",""), "风险": s.get("risk",""),
                })
            pd.DataFrame(cp_rows).to_excel(wb, sheet_name="精筛池", index=False)

    # Sheet 9: 实时信号（精筛池 + 盘中监测）
    import importlib
    cp2_path = os.path.join("data_cache", "curated_pool.json")
    signal_rows = []
    if os.path.exists(cp2_path):
        try:
            from real_time import get_realtime_quote
            with open(cp2_path, encoding="utf-8") as f:
                cps = json.load(f).get("stocks", [])
            for s in cps:
                sym = s["symbol"]
                quote = get_realtime_quote(sym)
                if not quote: continue
                price = quote.get("price") or 0
                chg = quote.get("change_pct") or 0
                turnover = quote.get("turnover_rate") or 0
                # 加载历史判断信号
                hist = None
                for fname in [f"{sym}.csv", f"{sym.lstrip('0')}.csv"]:
                    hp = os.path.join("data_cache", fname)
                    if os.path.exists(hp):
                        hist = pd.read_csv(hp, index_col=0, parse_dates=True)
                        break
                signals = []
                if hist is not None and len(hist) >= 20:
                    h20 = hist["high"].iloc[-20:].max()
                    avg_vol_5d = hist["volume"].iloc[-6:-1].mean()
                    if avg_vol_5d > 0:
                        vol_ratio = hist["volume"].iloc[-1] / avg_vol_5d
                        if vol_ratio > 1.5: signals.append(f"放量{vol_ratio:.1f}x")
                    if hist["close"].iloc[-1] >= h20 * 0.98: signals.append("接近新高")
                    if abs(chg) > 5: signals.append(f"{'大涨' if chg>0 else '大跌'}{abs(chg):.0f}%")
                signal_rows.append({
                    "代码": sym, "名称": s.get("name",""), "现价": f"{price:.2f}",
                    "涨跌": f"{chg:+.2f}%", "换手": f"{turnover:.1f}%",
                    "信号": " | ".join(signals) if signals else "—",
                })
        except Exception:
            pass
    if signal_rows:
        pd.DataFrame(signal_rows).to_excel(wb, sheet_name="实时信号", index=False)

    wb.close()

    # ── 格式化 ──
    from openpyxl import load_workbook
    wb2 = load_workbook(path)
    for ws in wb2.worksheets:
        if ws.max_row > 1:
            style_header(ws, 1, ws.max_column)
        auto_width(ws)
    wb2.save(path)

    print(f"综合日报已保存: {path}")
    return path


if __name__ == "__main__":
    generate_report()
